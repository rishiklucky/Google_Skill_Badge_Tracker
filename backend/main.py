from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
import requests
from bs4 import BeautifulSoup
import re
import os
import uuid
import tempfile
import io
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Google Skill Badge Tracker")

# In production the React build is served by FastAPI itself,
# so CORS is only needed for local development.
ALLOWED_ORIGINS = os.environ.get(
    "ALLOWED_ORIGINS",
    "http://localhost:5173,http://localhost:3000"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMP_DIR = tempfile.gettempdir()

# Path to the Vite build output (relative to this file)
STATIC_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def normalize(name: str) -> str:
    """Lowercase + collapse whitespace."""
    return re.sub(r"\s+", " ", name.strip().lower())


def _looks_like_url(text: str) -> bool:
    return text.startswith("http://") or text.startswith("https://")


# ─────────────────────────────────────────────
#  Scraping
# ─────────────────────────────────────────────

def scrape_badges(profile_url: str) -> list[str]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
    }
    try:
        response = requests.get(profile_url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        badges = extract_badges_from_soup(soup)
        if badges:
            logger.info(f"Scraped {len(badges)} badges via requests.")
            return badges
        logger.info("No badges via requests, trying Selenium...")
    except Exception as e:
        logger.warning(f"requests scrape failed: {e}")
    return scrape_with_selenium(profile_url)


def extract_badges_from_soup(soup: BeautifulSoup) -> list[str]:
    badges = []
    selectors = [
        {"tag": "span", "class_": "ql-title-medium"},
        {"tag": "h3",   "class_": "profile-badge__name"},
        {"tag": "div",  "class_": "badge-title"},
        {"tag": "p",    "class_": "l-mts"},
    ]
    for sel in selectors:
        elements = soup.find_all(sel["tag"], class_=sel.get("class_"))
        if elements:
            for el in elements:
                text = el.get_text(strip=True)
                if text:
                    badges.append(text)
            if badges:
                return badges
    for el in soup.find_all(True):
        classes = " ".join(el.get("class", []))
        if "badge" in classes.lower() and el.name in ("span", "p", "h3", "h4", "div"):
            text = el.get_text(strip=True)
            if text and len(text) > 3:
                badges.append(text)
    return list(dict.fromkeys(badges))


def scrape_with_selenium(profile_url: str) -> list[str]:
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        options = Options()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        driver = webdriver.Chrome(options=options)
        driver.get(profile_url)
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "span.ql-title-medium, .profile-badge__name, .badge-title")
                )
            )
        except Exception:
            import time
            time.sleep(5)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        driver.quit()
        badges = extract_badges_from_soup(soup)
        logger.info(f"Scraped {len(badges)} badges via Selenium.")
        return badges
    except ImportError:
        raise HTTPException(status_code=500, detail="Selenium not installed. Run: pip install selenium")
    except Exception as e:
        logger.error(f"Selenium scrape failed: {e}")
        raise HTTPException(status_code=500, detail=f"Scraping failed: {str(e)}")


# ─────────────────────────────────────────────
#  Excel read  (preserves hyperlinks)
# ─────────────────────────────────────────────

def read_xlsx(file_bytes: bytes):
    """
    Returns (col_headers, data_rows) where each data_row is a dict:
      {
        "badge_name":  str,   # human-readable name
        "availability": str,  # column-2 text
        "hyperlink":   str|None,  # URL from column-1 hyperlink
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=False))
    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file is empty or has no data rows.")

    # ── header row ──────────────────────────────────────────────────────────
    header_row = all_rows[0]
    col1_header = str(header_row[0].value).strip() if header_row[0].value else "Badge Name"
    col2_header = str(header_row[1].value).strip() if len(header_row) > 1 and header_row[1].value else "Availability Status"

    # ── data rows ────────────────────────────────────────────────────────────
    data_rows = []
    for raw_row in all_rows[1:]:
        # ── Column 1: badge name + hyperlink ────────────────────────────────
        cell1 = raw_row[0] if len(raw_row) > 0 else None
        raw_val = str(cell1.value).strip() if (cell1 and cell1.value is not None) else ""

        hl = cell1.hyperlink if cell1 else None
        hl_target  = (hl.target  or "").strip() if hl else ""
        hl_display = (hl.display or "").strip() if hl else ""

        # Resolve the true badge name
        if hl_display and not _looks_like_url(hl_display):
            badge_name = hl_display
            link_url   = hl_target or (raw_val if _looks_like_url(raw_val) else None)
        elif raw_val and not _looks_like_url(raw_val):
            badge_name = raw_val
            link_url   = hl_target or None
        else:
            # cell value is a URL — use display or raw_val as name
            badge_name = hl_display or raw_val
            link_url   = hl_target or raw_val or None

        # ── Column 2: availability ───────────────────────────────────────────
        cell2 = raw_row[1] if len(raw_row) > 1 else None
        availability = str(cell2.value).strip() if (cell2 and cell2.value is not None) else ""

        # Skip fully blank rows
        if not badge_name:
            continue

        data_rows.append({
            "badge_name":   badge_name,
            "availability": availability,
            "hyperlink":    link_url,
        })

    return col1_header, col2_header, data_rows


# ─────────────────────────────────────────────
#  Excel write  (with Completion Status column)
# ─────────────────────────────────────────────

def write_xlsx(col1_header: str, col2_header: str, data_rows: list[dict]) -> bytes:
    """
    Writes three columns:
      1. Badge Name  (with hyperlinks restored)
      2. Availability Status
      3. Completion Status  — Excel conditional formatting rules so colour
         updates automatically when you edit the cell manually.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Badge Tracker"

    num_rows  = len(data_rows)
    last_row  = max(num_rows + 1, 2)
    # Apply CF to a large range so manually added rows also get formatting
    cf_range  = f"C2:C{last_row + 500}"

    # ── header row ───────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="4285F4")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate([col1_header, col2_header, "Completion Status"], start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    # ── data rows ─────────────────────────────────────────────────────────────
    for r, row in enumerate(data_rows, start=2):
        badge_name        = row["badge_name"]
        availability      = row["availability"]
        completion_status = row["completion_status"]
        link_url          = row.get("hyperlink")

        c1 = ws.cell(row=r, column=1, value=badge_name)
        c1.alignment = Alignment(vertical="center")
        if link_url:
            c1.hyperlink = str(link_url)
            c1.font = Font(color="1155CC", underline="single")

        c2 = ws.cell(row=r, column=2, value=availability)
        c2.alignment = Alignment(vertical="center")

        c3 = ws.cell(row=r, column=3, value=completion_status)
        c3.alignment = Alignment(horizontal="center", vertical="center")

    # ── conditional formatting (formula-based, Excel-native) ─────────────────
    #
    # openpyxl's "containsText" rule type omits the required <formula> child
    # that Excel needs, so we use type="formula" directly — exactly what Excel
    # writes when you create a "text contains" rule through the UI.
    #
    # SEARCH() is case-insensitive. C2 is the anchor cell of the range.
    # Priority order matters: "Not Completed" must have HIGHER priority (lower
    # number) than "Completed" so it wins when the cell says "Not Completed"
    # (which also contains the substring "Completed").

    def make_rule(formula: str, bg: str, fg: str, priority: int) -> Rule:
        dxf = DifferentialStyle(
            fill=PatternFill(
                patternType="solid",
                fgColor=f"FF{bg}",   # FF prefix = fully opaque
                bgColor=f"FF{bg}",
            ),
            font=Font(bold=True, color=f"FF{fg}"),
        )
        return Rule(type="expression", formula=[formula], dxf=dxf, priority=priority)

    # Priority 1 = highest (evaluated first by Excel)
    # "Not Completed" must beat "Completed" since "Completed" is a substring
    ws.conditional_formatting.add(
        cf_range,
        make_rule(formula='EXACT(C2,"Not Completed")', bg="FEF7E0", fg="B06000", priority=1),
    )
    ws.conditional_formatting.add(
        cf_range,
        make_rule(formula='EXACT(C2,"Completed")',     bg="E6F4EA", fg="137333", priority=2),
    )
    ws.conditional_formatting.add(
        cf_range,
        make_rule(formula='EXACT(C2,"Not Available")', bg="FCE8E6", fg="C5221F", priority=3),
    )

    # ── column widths ─────────────────────────────────────────────────────────
    max1 = max((len(r["badge_name"]) for r in data_rows), default=10)
    ws.column_dimensions["A"].width = min(max1 + 4, 60)
    max2 = max((len(r["availability"]) for r in data_rows), default=15)
    ws.column_dimensions["B"].width = min(max2 + 4, 30)
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
#  Core processing logic
# ─────────────────────────────────────────────

def compute_completion(data_rows: list[dict], completed_badges: list[str]) -> tuple[list[dict], dict]:
    """
    Adds 'completion_status' key to every row dict.
    Returns (updated_rows, summary).
    """
    completed_set = {normalize(b) for b in completed_badges}

    for row in data_rows:
        avail = row["availability"].strip().lower()
        name  = normalize(row["badge_name"])

        if avail in ("not available", "notavailable", "unavailable", "no"):
            row["completion_status"] = "Not Available"
        elif name in completed_set:
            row["completion_status"] = "Completed"
        else:
            row["completion_status"] = "Not Completed"

    summary = {
        "total":         len(data_rows),
        "completed":     sum(1 for r in data_rows if r["completion_status"] == "Completed"),
        "not_completed": sum(1 for r in data_rows if r["completion_status"] == "Not Completed"),
        "not_available": sum(1 for r in data_rows if r["completion_status"] == "Not Available"),
    }

    logger.info(f"Summary: {summary}")
    return data_rows, summary


# ─────────────────────────────────────────────
#  API endpoints
# ─────────────────────────────────────────────

@app.post("/api/process")
async def process(
    profile_url: str = Form(...),
    file: UploadFile = File(...),
):
    if not profile_url.startswith("http"):
        raise HTTPException(status_code=400, detail="Invalid URL.")

    file_bytes = await file.read()
    filename   = file.filename or "upload.xlsx"

    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are supported.")

    # 1. Scrape
    logger.info(f"Scraping: {profile_url}")
    completed_badges = scrape_badges(profile_url)
    if not completed_badges:
        raise HTTPException(
            status_code=422,
            detail="No badges found on the profile. Make sure your profile is public.",
        )

    # 2. Read Excel
    try:
        col1_header, col2_header, data_rows = read_xlsx(file_bytes)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    # 3. Compute completion status
    data_rows, summary = compute_completion(data_rows, completed_badges)

    # 4. Write output xlsx
    out_bytes = write_xlsx(col1_header, col2_header, data_rows)

    # 5. Save to temp
    file_id  = str(uuid.uuid4())
    out_path = os.path.join(TEMP_DIR, f"{file_id}.xlsx")
    with open(out_path, "wb") as f:
        f.write(out_bytes)

    logger.info(f"Saved output to {out_path} ({len(out_bytes)} bytes)")

    not_completed_badges = [
        r["badge_name"] for r in data_rows if r["completion_status"] == "Not Completed"
    ]

    return {
        "file_id":             file_id,
        "summary":             summary,
        "scraped_badges":      completed_badges,
        "not_completed_badges": not_completed_badges,
    }


@app.post("/api/verify")
async def verify(file: UploadFile = File(...)):
    """
    Upload a previously edited xlsx and get a count of each unique value
    in the Completion Status column (column 3).
    Works with any values — including ones the user typed manually.
    """
    file_bytes = await file.read()
    filename   = file.filename or "upload.xlsx"

    try:
        if filename.endswith(".csv"):
            import pandas as pd
            df = pd.read_csv(io.BytesIO(file_bytes), keep_default_na=False, dtype=str)
            # Find the Completion Status column (3rd col or by name)
            status_col = None
            for col in df.columns:
                if "completion" in col.lower() or "status" in col.lower():
                    status_col = col
                    break
            if status_col is None and df.shape[1] >= 3:
                status_col = df.columns[2]
            if status_col is None:
                raise HTTPException(status_code=400, detail="Could not find Completion Status column.")
            values = df[status_col].astype(str).str.strip().tolist()
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                raise HTTPException(status_code=400, detail="File has no data rows.")

            # Find Completion Status column index from header row
            header = [str(c).strip().lower() if c else "" for c in all_rows[0]]
            col_idx = None
            for i, h in enumerate(header):
                if "completion" in h or ("status" in h and i >= 2):
                    col_idx = i
                    break
            # Fallback: use 3rd column (index 2)
            if col_idx is None:
                col_idx = 2

            values = []
            for row in all_rows[1:]:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    values.append(str(val).strip())

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    # Count each unique value, skip blanks
    from collections import Counter
    counts = Counter(v for v in values if v)

    # Build response — sort by count descending
    breakdown = [
        {"label": label, "count": count}
        for label, count in sorted(counts.items(), key=lambda x: -x[1])
    ]

    total = sum(c["count"] for c in breakdown)

    return {
        "total": total,
        "breakdown": breakdown,
    }


@app.get("/api/download/{file_id}")
def download(file_id: str):
    if not re.match(r"^[a-f0-9\-]{36}$", file_id):
        raise HTTPException(status_code=400, detail="Invalid file ID.")

    path = os.path.join(TEMP_DIR, f"{file_id}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found or expired.")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="badge_tracker_result.xlsx",
    )


@app.post("/api/inspect")
async def inspect_file(file: UploadFile = File(...)):
    """Debug: see exactly what the backend reads from your Excel file."""
    file_bytes = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=False))
    result = []
    for i, row in enumerate(all_rows[:11]):
        row_info = []
        for j, cell in enumerate(row[:3]):
            hl = cell.hyperlink
            row_info.append({
                "col":        j,
                "value":      str(cell.value) if cell.value is not None else None,
                "hl_target":  (hl.target  or None) if hl else None,
                "hl_display": (hl.display or None) if hl else None,
            })
        result.append({"row": i, "cells": row_info})
    return {"rows": result}


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/template")
def download_template():
    """
    Serve the badge list template xlsx with the correct MIME type.
    The file lives in frontend/public/badge_list.xlsx (copied to dist/ at build time).
    """
    # Look in the built dist folder first (production), then fall back to public/ (dev)
    candidates = [
        os.path.join(os.path.dirname(__file__), "..", "frontend", "dist",    "badge_list.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "frontend", "public",  "badge_list.xlsx"),
    ]
    path = next((p for p in candidates if os.path.isfile(p)), None)

    if not path:
        raise HTTPException(status_code=404, detail="Template file not found on server.")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="Google_Skill_Badge_List.xlsx",
    )


# ─────────────────────────────────────────────
#  Serve React frontend (production)
#  Must be registered AFTER all /api routes.
# ─────────────────────────────────────────────

if os.path.isdir(STATIC_DIR):
    # Serve static assets (JS, CSS, images)
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def serve_spa(full_path: str):
        """Catch-all: return index.html so React Router handles client-side routing."""
        index = os.path.join(STATIC_DIR, "index.html")
        return FileResponse(index)
else:
    logger.warning(
        f"Frontend build not found at {STATIC_DIR}. "
        "Run 'npm run build' inside the frontend/ directory."
    )
